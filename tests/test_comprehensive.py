"""
Comprehensive test suite for the Manufacturing QC Automation System.
Covers: DataCleaner, Validator, QualityEvaluator, ReportGenerator,
        BatchManager, InputHandler, ExcelParser, and end-to-end integration.
"""
import pytest
import math
import os
import shutil
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock

import sys
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# Module imports
from data_cleaner import DataCleaner
from validator import Validator, ValidationError
from quality_rules import QualityEvaluator
from report_generator import ReportGenerator
from batch_manager import BatchManager
from input_handler import InputHandler
from config import (
    BUS_BARS, POINTS_PER_BAR, TOTAL_POINTS,
    RULE_A_THRESHOLD, MIN_POINTS_RULE_A, RULE_A_PERCENTAGE,
    RULE_B_THRESHOLD, MAX_RULE_B_PER_BAR,
    RULE_C_THRESHOLD, MAX_RULE_C_TOTAL, MAX_RULE_C_PER_BAR,
    DATA_VALUE_MIN, DATA_VALUE_MAX,
    CATEGORY_APPROVED, CATEGORY_REJECTED, CATEGORY_DATA_ERROR,
    MAX_FILE_SIZE_BYTES,
    OUTPUT_DIR,
)


# ===== HELPERS =====
def make_matrix(val=1.0):
    """Generate a 16x7 matrix filled with `val`."""
    return [[val for _ in range(POINTS_PER_BAR)] for _ in range(BUS_BARS)]


# ===================================================================
#  DATA CLEANER TESTS
# ===================================================================
class TestDataCleaner:
    # --- clean_value ---
    def test_float_passthrough(self):
        assert DataCleaner.clean_value(1.5) == 1.5

    def test_int_passthrough(self):
        assert DataCleaner.clean_value(2) == 2.0

    def test_string_normal(self):
        assert DataCleaner.clean_value("0.85") == 0.85

    def test_ocr_uppercase_O(self):
        assert DataCleaner.clean_value("O.85") == 0.85

    def test_ocr_lowercase_o(self):
        assert DataCleaner.clean_value("o.85") == 0.85

    def test_ocr_comma_decimal(self):
        assert DataCleaner.clean_value("0,85") == 0.85

    def test_whitespace_stripping(self):
        assert DataCleaner.clean_value("  0.90  ") == 0.90

    def test_combined_ocr_errors(self):
        # "O,85" => replace O->0, comma->. => "0.85"
        assert DataCleaner.clean_value("O,85") == 0.85

    def test_garbage_raises_valueerror(self):
        with pytest.raises(ValueError):
            DataCleaner.clean_value("X.XX")

    def test_empty_string_raises(self):
        with pytest.raises(ValueError):
            DataCleaner.clean_value("")

    def test_non_string_non_numeric_raises(self):
        with pytest.raises(ValueError):
            DataCleaner.clean_value([1, 2, 3])

    def test_none_raises(self):
        with pytest.raises((ValueError, TypeError)):
            DataCleaner.clean_value(None)

    def test_negative_value_clamped_to_min(self):
        result = DataCleaner.clean_value("-1.5")
        assert result == DATA_VALUE_MIN

    def test_extreme_high_value_clamped_to_max(self):
        result = DataCleaner.clean_value("99.9")
        assert result == DATA_VALUE_MAX

    def test_boundary_value_at_max(self):
        result = DataCleaner.clean_value("10.0")
        assert result == 10.0

    def test_boundary_value_at_min(self):
        result = DataCleaner.clean_value("0.0")
        assert result == 0.0

    def test_zero(self):
        assert DataCleaner.clean_value("0") == 0.0

    # --- clean_matrix ---
    def test_clean_matrix_normal(self):
        raw = [["0.85", "0,90", "O.75"] for _ in range(3)]
        cleaned = DataCleaner.clean_matrix(raw)
        assert len(cleaned) == 3
        assert all(isinstance(v, float) for row in cleaned for v in row)

    def test_clean_matrix_mixed_types(self):
        raw = [[0.5, "0.6", 1]]
        cleaned = DataCleaner.clean_matrix(raw)
        assert cleaned == [[0.5, 0.6, 1.0]]

    def test_clean_matrix_raises_on_garbage_cell(self):
        raw = [["0.5", "GARBAGE", "0.6"]]
        with pytest.raises(ValueError):
            DataCleaner.clean_matrix(raw)

    def test_clean_matrix_empty(self):
        assert DataCleaner.clean_matrix([]) == []


# ===================================================================
#  VALIDATOR TESTS
# ===================================================================
class TestValidator:
    def test_perfect_16x7_passes(self):
        Validator.validate_matrix(make_matrix(1.0))

    def test_too_few_rows(self):
        m = make_matrix()
        m.pop()  # 15 rows — above MIN_BUS_BARS, returns warning
        result = Validator.validate_matrix(m)
        assert result is not None  # ValidationWarning

    def test_too_many_rows(self):
        m = make_matrix()
        m.append([1.0] * POINTS_PER_BAR)  # 17 rows — above BUS_BARS, returns warning
        result = Validator.validate_matrix(m)
        assert result is not None  # ValidationWarning

    def test_too_few_cols(self):
        m = make_matrix()
        m[5] = [1.0] * 6  # 6 cols on row 5
        with pytest.raises(ValidationError, match="got 6"):
            Validator.validate_matrix(m)

    def test_too_many_cols(self):
        m = make_matrix()
        m[0].append(1.0)  # 8 cols on row 0
        with pytest.raises(ValidationError, match="got 8"):
            Validator.validate_matrix(m)

    def test_none_value_detected(self):
        m = make_matrix()
        m[3][4] = None
        with pytest.raises(ValidationError):
            Validator.validate_matrix(m)

    def test_nan_value_detected(self):
        m = make_matrix()
        m[0][0] = float('nan')
        with pytest.raises(ValidationError):
            Validator.validate_matrix(m)

    def test_string_value_rejected(self):
        m = make_matrix()
        m[2][2] = "hello"
        with pytest.raises(ValidationError, match="Non-numeric"):
            Validator.validate_matrix(m)

    def test_empty_matrix_fails(self):
        with pytest.raises(ValidationError):
            Validator.validate_matrix([])

    def test_single_row_fails(self):
        with pytest.raises(ValidationError):
            Validator.validate_matrix([[1.0] * 7])

    def test_all_zeros_passes(self):
        """All zeros are still numeric: structural validation should pass."""
        Validator.validate_matrix(make_matrix(0.0))

    def test_negative_floats_pass_structurally(self):
        """Validator checks structure only, not value ranges."""
        m = make_matrix(-1.0)
        Validator.validate_matrix(m)  # Should pass structural check


# ===================================================================
#  QUALITY RULES TESTS
# ===================================================================
class TestRuleA:
    def test_all_above_threshold_passes(self):
        m = make_matrix(1.0)
        passed, count, required = QualityEvaluator.evaluate_rule_a(m)
        assert passed is True
        assert count == TOTAL_POINTS

    def test_all_below_threshold_fails(self):
        m = make_matrix(0.5)
        passed, count, required = QualityEvaluator.evaluate_rule_a(m)
        assert passed is False
        assert count == 0

    def test_exact_threshold_boundary_excluded(self):
        """Points exactly at 0.8 should NOT count (must be > 0.8)."""
        m = make_matrix(0.8)
        passed, count, required = QualityEvaluator.evaluate_rule_a(m)
        assert count == 0
        assert passed is False

    def test_exact_84_passes(self):
        m = make_matrix(0.5)
        for i in range(12):
            for j in range(7):
                m[i][j] = 0.9
        passed, count, required = QualityEvaluator.evaluate_rule_a(m)
        assert count == 84
        assert passed is True

    def test_83_fails(self):
        m = make_matrix(0.5)
        for i in range(12):
            for j in range(7):
                m[i][j] = 0.9
        m[0][0] = 0.5
        passed, count, required = QualityEvaluator.evaluate_rule_a(m)
        assert count == 83
        assert passed is False

    def test_min_points_config_is_correct(self):
        assert MIN_POINTS_RULE_A == 84
        assert RULE_A_PERCENTAGE == 0.75


class TestRuleB:
    def test_all_good_passes(self):
        m = make_matrix(0.9)
        passed, fb = QualityEvaluator.evaluate_rule_b(m)
        assert passed is True
        assert all(v == 0 for v in fb.values())

    def test_exactly_2_per_bar_passes(self):
        m = make_matrix(0.9)
        m[0][0] = 0.3
        m[0][1] = 0.3
        passed, fb = QualityEvaluator.evaluate_rule_b(m)
        assert passed is True
        assert fb[0] == 2

    def test_3_per_bar_fails(self):
        m = make_matrix(0.9)
        m[0][0] = 0.3
        m[0][1] = 0.3
        m[0][2] = 0.3
        passed, fb = QualityEvaluator.evaluate_rule_b(m)
        assert passed is False
        assert fb[0] == 3

    def test_exactly_at_threshold_counts_as_failure(self):
        """Values exactly at 0.35 should count as ≤0.35 failures."""
        m = make_matrix(0.9)
        m[0][0] = 0.35
        m[0][1] = 0.35
        m[0][2] = 0.35
        passed, fb = QualityEvaluator.evaluate_rule_b(m)
        assert fb[0] == 3
        assert passed is False

    def test_multiple_bars_failing(self):
        m = make_matrix(0.9)
        for bar in range(4):
            m[bar][0] = 0.2
            m[bar][1] = 0.2
            m[bar][2] = 0.2
        passed, fb = QualityEvaluator.evaluate_rule_b(m)
        assert passed is False
        for bar in range(4):
            assert fb[bar] == 3

    def test_failures_spread_across_bars_ok(self):
        """2 failures in each bar is fine, even if total is high."""
        m = make_matrix(0.9)
        for bar in range(BUS_BARS):
            m[bar][0] = 0.3
            m[bar][1] = 0.3
        passed, _ = QualityEvaluator.evaluate_rule_b(m)
        assert passed is True


class TestRuleC:
    def test_all_good_passes(self):
        m = make_matrix(0.9)
        passed, total, pb = QualityEvaluator.evaluate_rule_c(m)
        assert passed is True
        assert total == 0

    def test_8_failures_across_8_bars_passes(self):
        m = make_matrix(0.9)
        for i in range(8):
            m[i][0] = 0.05
        passed, total, _ = QualityEvaluator.evaluate_rule_c(m)
        assert passed is True
        assert total == 8

    def test_9_failures_total_fails(self):
        m = make_matrix(0.9)
        for i in range(9):
            m[i][0] = 0.05
        passed, total, _ = QualityEvaluator.evaluate_rule_c(m)
        assert passed is False
        assert total == 9

    def test_2_per_bar_fails_even_if_total_low(self):
        m = make_matrix(0.9)
        m[0][0] = 0.05
        m[0][1] = 0.05
        passed, total, _ = QualityEvaluator.evaluate_rule_c(m)
        assert passed is False
        assert total == 2

    def test_exactly_at_threshold_counts(self):
        """Values exactly at 0.1 should count as ≤0.1 failures."""
        m = make_matrix(0.9)
        m[0][0] = 0.1
        passed, total, pb = QualityEvaluator.evaluate_rule_c(m)
        assert pb[0] == 1
        assert total == 1

    def test_zero_value_counts(self):
        m = make_matrix(0.9)
        m[0][0] = 0.0
        passed, total, _ = QualityEvaluator.evaluate_rule_c(m)
        assert total == 1


class TestEvaluateBatch:
    def test_perfect_batch_approved(self):
        m = make_matrix(1.0)
        report = QualityEvaluator.evaluate_batch(m)
        assert report["decision"] == CATEGORY_APPROVED
        assert report["metrics"]["rule_A"]["passed"] is True
        assert report["metrics"]["rule_B"]["passed"] is True
        assert report["metrics"]["rule_C"]["passed"] is True

    def test_all_bad_rejected(self):
        m = make_matrix(0.05)
        report = QualityEvaluator.evaluate_batch(m)
        assert report["decision"] == CATEGORY_REJECTED

    def test_only_rule_a_fails_rejected(self):
        """Rule B and C pass, but A fails => still REJECTED."""
        m = make_matrix(0.5)  # All 0.5: A fails (0 > 0.8), B and C pass
        report = QualityEvaluator.evaluate_batch(m)
        assert report["decision"] == CATEGORY_REJECTED
        assert report["metrics"]["rule_A"]["passed"] is False
        assert report["metrics"]["rule_B"]["passed"] is True
        assert report["metrics"]["rule_C"]["passed"] is True

    def test_only_rule_b_fails_rejected(self):
        m = make_matrix(0.9)
        # 3 failures ≤0.35 in bar 0 => B fails
        m[0][0] = 0.3
        m[0][1] = 0.3
        m[0][2] = 0.3
        report = QualityEvaluator.evaluate_batch(m)
        assert report["decision"] == CATEGORY_REJECTED
        assert report["metrics"]["rule_B"]["passed"] is False

    def test_only_rule_c_fails_rejected(self):
        m = make_matrix(0.9)
        # 9 failures ≤0.1 in 9 bars => C fails (total > 8)
        for i in range(9):
            m[i][0] = 0.05
        report = QualityEvaluator.evaluate_batch(m)
        assert report["decision"] == CATEGORY_REJECTED
        assert report["metrics"]["rule_C"]["passed"] is False

    def test_report_has_required_keys(self):
        m = make_matrix(1.0)
        report = QualityEvaluator.evaluate_batch(m)
        assert "decision" in report
        assert "metrics" in report
        assert "rule_A" in report["metrics"]
        assert "rule_B" in report["metrics"]
        assert "rule_C" in report["metrics"]


# ===================================================================
#  BATCH MANAGER TESTS
# ===================================================================
class TestBatchManager:
    def test_batch_id_format(self):
        bm = BatchManager()
        assert bm.batch_id.startswith("BATCH_")
        parts = bm.batch_id.split("_")
        assert len(parts) >= 3

    def test_batch_ids_are_unique(self):
        ids = {BatchManager().batch_id for _ in range(20)}
        assert len(ids) == 20

    def test_default_operator_and_machine(self):
        bm = BatchManager()
        assert bm.operator_name == "AUTO_SYSTEM"
        assert bm.machine_id == "SYS_01"

    def test_custom_operator_and_machine(self):
        bm = BatchManager(operator_name="Bob", machine_id="M42")
        assert bm.operator_name == "Bob"
        assert bm.machine_id == "M42"

    def test_set_and_get_metadata(self):
        bm = BatchManager()
        bm.set_metadata("shift", "B")
        bm.set_metadata("line", 3)
        ctx = bm.get_context()
        assert ctx["metadata"]["shift"] == "B"
        assert ctx["metadata"]["line"] == 3

    def test_get_context_structure(self):
        bm = BatchManager()
        ctx = bm.get_context()
        assert "batch_id" in ctx
        assert "timestamp" in ctx
        assert "operator" in ctx
        assert "machine_id" in ctx
        assert "metadata" in ctx

    def test_log_context_does_not_crash(self):
        bm = BatchManager()
        bm.log_context("Test message")  # Should not raise


# ===================================================================
#  REPORT GENERATOR TESTS
# ===================================================================
class TestReportGenerator:
    def test_generate_approved_report(self, tmp_path):
        m = make_matrix(1.0)
        report = QualityEvaluator.evaluate_batch(m)
        with patch.object(ReportGenerator, '_get_output_dir', return_value=tmp_path, create=True):
            # Patch OUTPUT_DIR for the test
            import report_generator
            original_dir = report_generator.OUTPUT_DIR
            report_generator.OUTPUT_DIR = tmp_path
            try:
                path = ReportGenerator.generate_report("TEST_BATCH_001", m, report)
                assert path.exists()
                assert path.suffix == ".xlsx"
                assert "TEST_BATCH_001" in path.name
            finally:
                report_generator.OUTPUT_DIR = original_dir

    def test_generate_rejected_report(self, tmp_path):
        m = make_matrix(0.05)
        report = QualityEvaluator.evaluate_batch(m)
        import report_generator
        original_dir = report_generator.OUTPUT_DIR
        report_generator.OUTPUT_DIR = tmp_path
        try:
            path = ReportGenerator.generate_report("TEST_BATCH_002", m, report)
            assert path.exists()
            assert "TEST_BATCH_002" in path.name
        finally:
            report_generator.OUTPUT_DIR = original_dir

    def test_empty_matrix_raises(self):
        with pytest.raises(ValueError, match="matrix is empty"):
            ReportGenerator.generate_report("X", [], {"decision": "REJECTED", "metrics": {}})

    def test_report_contains_both_sheets(self, tmp_path):
        import openpyxl
        import report_generator
        m = make_matrix(1.0)
        report = QualityEvaluator.evaluate_batch(m)
        original_dir = report_generator.OUTPUT_DIR
        report_generator.OUTPUT_DIR = tmp_path
        try:
            path = ReportGenerator.generate_report("TEST_SHEETS", m, report)
            wb = openpyxl.load_workbook(path)
            assert "Cleaned Data" in wb.sheetnames
            assert "QC Summary" in wb.sheetnames
        finally:
            report_generator.OUTPUT_DIR = original_dir


# ===================================================================
#  INPUT HANDLER TESTS
# ===================================================================
class TestInputHandler:
    def test_supported_extensions(self):
        assert '.xlsx' in InputHandler.SUPPORTED_EXTENSIONS
        assert '.png' in InputHandler.SUPPORTED_EXTENSIONS
        assert '.jpg' in InputHandler.SUPPORTED_EXTENSIONS
        assert '.jpeg' in InputHandler.SUPPORTED_EXTENSIONS

    def test_unsupported_extension_returns_data_error(self, tmp_path):
        f = tmp_path / "test.pdf"
        f.write_text("fake")
        bm = BatchManager()
        matrix, cat = InputHandler.route_file(f, bm)
        assert matrix is None
        assert cat == CATEGORY_DATA_ERROR

    def test_oversized_file_returns_data_error(self, tmp_path):
        f = tmp_path / "huge.xlsx"
        # Write a file just over the limit
        f.write_bytes(b'\x00' * (MAX_FILE_SIZE_BYTES + 1))
        bm = BatchManager()
        matrix, cat = InputHandler.route_file(f, bm)
        assert matrix is None
        assert cat == CATEGORY_DATA_ERROR

    def test_sanitize_filename_strips_path(self):
        result = InputHandler._sanitize_filename("../../etc/passwd")
        assert "/" not in result
        assert "\\" not in result
        assert "passwd" in result

    def test_sanitize_filename_safe_chars(self):
        result = InputHandler._sanitize_filename("my file (1).xlsx")
        assert "my_file_(1).xlsx" == result

    def test_move_file_to_processed(self, tmp_path):
        src = tmp_path / "test.xlsx"
        src.write_text("data")
        processed = tmp_path / "processed"
        processed.mkdir()
        with patch('input_handler.PROCESSED_DIR', processed), \
             patch('input_handler.FAILED_DIR', tmp_path / "failed"):
            InputHandler.move_file(src, success=True, batch_id="BATCH_TEST")
            assert not src.exists()
            moved = list(processed.iterdir())
            assert len(moved) == 1
            assert "BATCH_TEST" in moved[0].name

    def test_move_file_to_failed(self, tmp_path):
        src = tmp_path / "bad.xlsx"
        src.write_text("data")
        failed = tmp_path / "failed"
        failed.mkdir()
        with patch('input_handler.FAILED_DIR', failed), \
             patch('input_handler.PROCESSED_DIR', tmp_path / "processed"):
            InputHandler.move_file(src, success=False, batch_id="BATCH_FAIL")
            assert not src.exists()
            moved = list(failed.iterdir())
            assert len(moved) == 1
            assert "BATCH_FAIL" in moved[0].name


# ===================================================================
#  EXCEL PARSER TESTS
# ===================================================================
class TestExcelParser:
    def _make_test_excel(self, tmp_path, matrix, filename="test.xlsx"):
        """Helper to create a test xlsx file from a 2D list."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        for r_idx, row in enumerate(matrix, start=1):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        path = tmp_path / filename
        wb.save(path)
        return path

    def test_parse_valid_16x7(self, tmp_path):
        from excel_parser import ExcelParser
        raw = make_matrix(0.9)
        path = self._make_test_excel(tmp_path, raw)
        result = ExcelParser.extract_matrix(str(path))
        assert len(result) == BUS_BARS
        assert all(len(row) == POINTS_PER_BAR for row in result)

    def test_parse_fewer_rows_raises(self, tmp_path):
        from excel_parser import ExcelParser
        raw = [[0.9] * POINTS_PER_BAR for _ in range(10)]  # Only 10 rows
        path = self._make_test_excel(tmp_path, raw)
        with pytest.raises(ValidationError):
            ExcelParser.extract_matrix(str(path))

    def test_parse_nonexistent_file_raises(self):
        from excel_parser import ExcelParser
        with pytest.raises(ValidationError):
            ExcelParser.extract_matrix("nonexistent_file.xlsx")

    def test_parse_with_string_numbers(self, tmp_path):
        from excel_parser import ExcelParser
        raw = [["0.9"] * POINTS_PER_BAR for _ in range(BUS_BARS)]
        path = self._make_test_excel(tmp_path, raw)
        result = ExcelParser.extract_matrix(str(path))
        assert all(isinstance(v, float) for row in result for v in row)


# ===================================================================
#  INTEGRATION / END-TO-END TESTS
# ===================================================================
class TestIntegration:
    def test_full_pipeline_approved(self, tmp_path):
        """End-to-end: create Excel -> parse -> validate -> evaluate -> report."""
        import openpyxl
        from excel_parser import ExcelParser
        import report_generator

        # 1. Create a clean Excel file
        m = make_matrix(1.0)
        wb = openpyxl.Workbook()
        ws = wb.active
        for r, row in enumerate(m, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        xlsx_path = tmp_path / "approved_test.xlsx"
        wb.save(xlsx_path)

        # 2. Parse
        matrix = ExcelParser.extract_matrix(str(xlsx_path))

        # 3. Validate
        Validator.validate_matrix(matrix)

        # 4. Evaluate
        report = QualityEvaluator.evaluate_batch(matrix)
        assert report["decision"] == CATEGORY_APPROVED

        # 5. Generate report
        original_dir = report_generator.OUTPUT_DIR
        report_generator.OUTPUT_DIR = tmp_path
        try:
            report_path = ReportGenerator.generate_report("INT_TEST_001", matrix, report)
            assert report_path.exists()
        finally:
            report_generator.OUTPUT_DIR = original_dir

    def test_full_pipeline_rejected(self, tmp_path):
        """End-to-end: failing data goes through full pipeline."""
        import openpyxl
        from excel_parser import ExcelParser
        import report_generator

        # Matrix with lots of low values — will fail all rules
        m = make_matrix(0.05)
        wb = openpyxl.Workbook()
        ws = wb.active
        for r, row in enumerate(m, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        xlsx_path = tmp_path / "rejected_test.xlsx"
        wb.save(xlsx_path)

        matrix = ExcelParser.extract_matrix(str(xlsx_path))
        Validator.validate_matrix(matrix)
        report = QualityEvaluator.evaluate_batch(matrix)
        assert report["decision"] == CATEGORY_REJECTED

        original_dir = report_generator.OUTPUT_DIR
        report_generator.OUTPUT_DIR = tmp_path
        try:
            report_path = ReportGenerator.generate_report("INT_TEST_002", matrix, report)
            assert report_path.exists()
        finally:
            report_generator.OUTPUT_DIR = original_dir

    def test_borderline_batch(self, tmp_path):
        """Borderline case: exactly 84 pass Rule A, but a Rule C failure flips to REJECTED."""
        import openpyxl
        from excel_parser import ExcelParser
        import report_generator

        # 12 rows of 0.9 (84 points > 0.8), 4 rows of 0.5 (good for B and C)
        m = [[0.9] * POINTS_PER_BAR for _ in range(12)] + \
            [[0.5] * POINTS_PER_BAR for _ in range(4)]

        # Inject 9 Rule C failures in separate bars => C fails
        for i in range(9):
            m[i][6] = 0.05

        wb = openpyxl.Workbook()
        ws = wb.active
        for r, row in enumerate(m, 1):
            for c, val in enumerate(row, 1):
                ws.cell(row=r, column=c, value=val)
        xlsx_path = tmp_path / "borderline_test.xlsx"
        wb.save(xlsx_path)

        matrix = ExcelParser.extract_matrix(str(xlsx_path))
        Validator.validate_matrix(matrix)
        report = QualityEvaluator.evaluate_batch(matrix)
        assert report["decision"] == CATEGORY_REJECTED
        assert report["metrics"]["rule_C"]["passed"] is False


# ===================================================================
#  CONFIG INVARIANT TESTS
# ===================================================================
class TestConfig:
    def test_total_points_equals_product(self):
        assert TOTAL_POINTS == BUS_BARS * POINTS_PER_BAR

    def test_min_points_rule_a_correct(self):
        assert MIN_POINTS_RULE_A == int(TOTAL_POINTS * RULE_A_PERCENTAGE)

    def test_thresholds_ordered(self):
        assert RULE_C_THRESHOLD < RULE_B_THRESHOLD < RULE_A_THRESHOLD

    def test_max_rule_c_per_bar_leq_max_b(self):
        assert MAX_RULE_C_PER_BAR <= MAX_RULE_B_PER_BAR

    def test_data_bounds(self):
        assert DATA_VALUE_MIN == 0.0
        assert DATA_VALUE_MAX == 10.0

    def test_category_constants(self):
        assert CATEGORY_APPROVED == "APPROVED"
        assert CATEGORY_REJECTED == "REJECTED"
