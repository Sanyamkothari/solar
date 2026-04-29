import os
import json
import sqlite3
import shutil
import tempfile
from pathlib import Path
from typing import Optional, List
from datetime import datetime

from fastapi import FastAPI, File, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from fastapi import Response, Request

# Adjust Python path
import sys
BASE_DIR = Path(__file__).parent
if str(BASE_DIR) not in sys.path:
    sys.path.append(str(BASE_DIR))

import config as cfg_module
from config import (
    INPUT_DIR, PROCESSED_DIR, FAILED_DIR, OUTPUT_DIR, LOGS_DIR,
    RAG_DB_PATH, ENABLE_RAG_CONTEXT, ENABLE_RAG_LLM_SUMMARY, ENABLE_OPERATOR_FEEDBACK,
)
from input_handler import InputHandler
from main import process_file
from rag_engine import BatchHistoryDB

app = FastAPI(title="Factory QC Web Dashboard")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

STATIC_DIR = BASE_DIR / "static"
STATIC_DIR.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

rag_history_db = BatchHistoryDB(RAG_DB_PATH)


def _file_info(f: Path) -> dict:
    stat = f.stat()
    return {
        "name": f.name,
        "size": stat.st_size,
        "size_display": _human_size(stat.st_size),
        "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
        "extension": f.suffix.lower(),
    }

def _human_size(num_bytes: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if num_bytes < 1024:
            return f"{num_bytes:.1f} {unit}"
        num_bytes /= 1024
    return f"{num_bytes:.1f} TB"


def _rag_summary(limit: int = 5) -> dict:
    """Summarize RAG history for the UI."""
    summary = {
        "enabled": ENABLE_RAG_CONTEXT,
        "llm_enabled": ENABLE_RAG_LLM_SUMMARY,
        "feedback_enabled": ENABLE_OPERATOR_FEEDBACK,
        "total_batches": 0,
        "feedback_cases": 0,
        "recent_batches": [],
        "recent_feedback_cases": [],
    }

    if not RAG_DB_PATH.exists():
        return summary

    try:
        with sqlite3.connect(RAG_DB_PATH) as conn:
            conn.row_factory = sqlite3.Row
            total_row = conn.execute("SELECT COUNT(*) AS total FROM batch_history").fetchone()
            feedback_row = conn.execute(
                """
                SELECT COUNT(*) AS total
                FROM batch_history
                WHERE COALESCE(root_cause, '') != '' OR COALESCE(operator_feedback, '') != ''
                """
            ).fetchone()
            recent_rows = conn.execute(
                """
                SELECT batch_id, timestamp, decision, shift, equipment_id,
                       root_cause, operator_feedback, feedback_confidence,
                       reviewed_by, reviewed_at, action_taken
                FROM batch_history
                ORDER BY created_at DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()
            feedback_rows = conn.execute(
                """
                SELECT batch_id, timestamp, decision, shift, equipment_id,
                       root_cause, operator_feedback, feedback_confidence,
                       reviewed_by, reviewed_at, action_taken
                FROM batch_history
                WHERE COALESCE(root_cause, '') != '' OR COALESCE(operator_feedback, '') != ''
                ORDER BY reviewed_at DESC, created_at DESC
                LIMIT ?
                """,
                (limit,),
            ).fetchall()

        summary["total_batches"] = int(total_row["total"] if total_row else 0)
        summary["feedback_cases"] = int(feedback_row["total"] if feedback_row else 0)
        summary["recent_batches"] = [dict(row) for row in recent_rows]
        summary["recent_feedback_cases"] = [dict(row) for row in feedback_rows]
    except Exception as e:
        summary["error"] = str(e)

    return summary


# ─────────────────────── Pages ───────────────────────
@app.get("/", response_class=HTMLResponse)
async def serve_index():
    with open(STATIC_DIR / "index.html", "r", encoding="utf-8") as f:
        html = f.read()

    css_ver = int((STATIC_DIR / "styles.css").stat().st_mtime) if (STATIC_DIR / "styles.css").exists() else 1
    js_ver = int((STATIC_DIR / "app.js").stat().st_mtime) if (STATIC_DIR / "app.js").exists() else 1

    html = html.replace('/static/styles.css', f'/static/styles.css?v={css_ver}')
    html = html.replace('/static/app.js', f'/static/app.js?v={js_ver}')
    return html


@app.get("/favicon.ico", include_in_schema=False)
async def favicon():
    favicon_path = STATIC_DIR / "favicon.ico"
    if favicon_path.exists() and favicon_path.is_file():
        return FileResponse(path=str(favicon_path), media_type="image/x-icon")
    return Response(status_code=204)


# ─────────────────────── Metrics ───────────────────────
@app.get("/api/metrics")
async def get_metrics():
    processed_count = len(list(PROCESSED_DIR.glob('*')))
    output_count = len(list(OUTPUT_DIR.glob('*')))
    pending_count = len(InputHandler.get_pending_files())
    rag_summary = _rag_summary(limit=1)
    return {
        "processed": processed_count,
        "reports": output_count,
        "pending": pending_count,
        "rag_cases": rag_summary.get("total_batches", 0),
        "feedback_cases": rag_summary.get("feedback_cases", 0),
    }


@app.get("/api/intelligence")
async def get_intelligence():
    return _rag_summary(limit=5)


# ─────────────────────── File Listings ───────────────────────
@app.get("/api/processed")
async def list_processed():
    files = sorted(PROCESSED_DIR.glob("*"), key=lambda f: f.stat().st_mtime, reverse=True)
    return {"files": [_file_info(f) for f in files if f.is_file()]}


@app.get("/api/failed")
async def list_failed():
    files = sorted(FAILED_DIR.glob("*"), key=lambda f: f.stat().st_mtime, reverse=True)
    return {"files": [_file_info(f) for f in files if f.is_file()]}


@app.get("/api/reports")
async def list_reports():
    files = sorted(OUTPUT_DIR.glob("*"), key=lambda f: f.stat().st_mtime, reverse=True)
    return {"files": [_file_info(f) for f in files if f.is_file()]}


@app.get("/api/reports/download/{filename}")
async def download_report(filename: str):
    filepath = OUTPUT_DIR / filename
    if not filepath.exists() or not filepath.is_file():
        return JSONResponse(status_code=404, content={"error": "File not found"})
    if not filepath.resolve().is_relative_to(OUTPUT_DIR.resolve()):
        return JSONResponse(status_code=403, content={"error": "Forbidden"})
    return FileResponse(path=str(filepath), filename=filename)


# ─────────────────────── Report Detail (for graphs) ───────────────────────
@app.get("/api/reports/detail/{filename}")
async def report_detail(filename: str):
    """Parse an Excel report and return structured data for graphing."""
    filepath = OUTPUT_DIR / filename
    if not filepath.exists() or not filepath.is_file():
        return JSONResponse(status_code=404, content={"error": "File not found"})
    if not filepath.resolve().is_relative_to(OUTPUT_DIR.resolve()):
        return JSONResponse(status_code=403, content={"error": "Forbidden"})

    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(filepath), data_only=True)

        # Extract data from "Cleaned Data" sheet
        ws = wb["Cleaned Data"]
        batch_id = str(ws["A1"].value or "")
        decision = str(ws["B2"].value or "UNKNOWN")

        # Read matrix (starts at row 6, col 2..8 for P1-P7, until no more data)
        matrix = []
        bar_labels = []
        row_idx = 6
        while True:
            label = ws.cell(row=row_idx, column=1).value
            if label is None:
                break
            bar_labels.append(str(label))
            row_vals = []
            for c in range(2, 9):
                v = ws.cell(row=row_idx, column=c).value
                row_vals.append(float(v) if v is not None else 0.0)
            matrix.append(row_vals)
            row_idx += 1

        # Extract QC Summary
        qc_summary = {}
        if "QC Summary" in wb.sheetnames:
            ws_qc = wb["QC Summary"]
            for r in range(2, 6):
                metric_name = ws_qc.cell(row=r, column=1).value
                metric_val = ws_qc.cell(row=r, column=2).value
                threshold = ws_qc.cell(row=r, column=3).value
                passed = ws_qc.cell(row=r, column=4).value
                if metric_name:
                    qc_summary[str(metric_name)] = {
                        "value": str(metric_val) if metric_val else "",
                        "threshold": str(threshold) if threshold else "",
                        "passed": str(passed) if passed else "",
                    }

        # Compute graph-ready stats
        all_values = [v for row in matrix for v in row if v >= 0]
        bar_averages = [sum(row) / max(len(row), 1) for row in matrix]
        point_averages = []
        if matrix:
            for col in range(len(matrix[0])):
                col_vals = [matrix[r][col] for r in range(len(matrix)) if matrix[r][col] >= 0]
                point_averages.append(sum(col_vals) / max(len(col_vals), 1))

        # Distribution bins
        bins = {"<0.1": 0, "0.1-0.35": 0, "0.35-0.8": 0, ">0.8": 0}
        for v in all_values:
            if v <= 0.1:
                bins["<0.1"] += 1
            elif v <= 0.35:
                bins["0.1-0.35"] += 1
            elif v <= 0.8:
                bins["0.35-0.8"] += 1
            else:
                bins[">0.8"] += 1

        return {
            "batch_id": batch_id,
            "decision": decision,
            "bar_labels": bar_labels,
            "matrix": matrix,
            "bar_averages": bar_averages,
            "point_averages": point_averages,
            "distribution": bins,
            "qc_summary": qc_summary,
            "stats": {
                "total_points": len(all_values),
                "mean": round(sum(all_values) / max(len(all_values), 1), 4),
                "min": round(min(all_values), 4) if all_values else 0,
                "max": round(max(all_values), 4) if all_values else 0,
            }
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={"error": str(e)})


@app.get("/api/processed/view/{filename}")
async def view_processed_file(filename: str):
    filepath = PROCESSED_DIR / filename
    if not filepath.exists() or not filepath.is_file():
        return JSONResponse(status_code=404, content={"error": "File not found"})
    if not filepath.resolve().is_relative_to(PROCESSED_DIR.resolve()):
        return JSONResponse(status_code=403, content={"error": "Forbidden"})
    ext = filepath.suffix.lower()
    media_types = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg'}
    media_type = media_types.get(ext, 'application/octet-stream')
    return FileResponse(path=str(filepath), media_type=media_type)


@app.get("/api/failed/view/{filename}")
async def view_failed_file(filename: str):
    filepath = FAILED_DIR / filename
    if not filepath.exists() or not filepath.is_file():
        return JSONResponse(status_code=404, content={"error": "File not found"})
    if not filepath.resolve().is_relative_to(FAILED_DIR.resolve()):
        return JSONResponse(status_code=403, content={"error": "Forbidden"})
    ext = filepath.suffix.lower()
    media_types = {'.png': 'image/png', '.jpg': 'image/jpeg', '.jpeg': 'image/jpeg'}
    media_type = media_types.get(ext, 'application/octet-stream')
    return FileResponse(path=str(filepath), media_type=media_type)


# ─────────────────────── Upload & Process ───────────────────────
@app.post("/api/upload")
async def upload_file(
    file: UploadFile = File(...),
    excel_ref: Optional[UploadFile] = File(None)
):
    try:
        tmp_dir = Path(tempfile.mkdtemp(prefix="qc_upload_"))
        file_path = tmp_dir / file.filename

        content = await file.read()
        with open(file_path, "wb") as f:
            f.write(content)

        excel_ref_path = None
        if excel_ref and excel_ref.filename:
            excel_ref_path = tmp_dir / f"ref_{excel_ref.filename}"
            excel_content = await excel_ref.read()
            with open(excel_ref_path, "wb") as f:
                f.write(excel_content)

        result = process_file(file_path, excel_ref_path=excel_ref_path)
        return JSONResponse(status_code=200, content={"success": True, "data": result})
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "error": str(e)})


# ─────────────────────── Logs ───────────────────────
@app.get("/api/logs")
async def get_logs():
    logs = list(LOGS_DIR.glob('*.log*'))
    if logs:
        latest_log = max(logs, key=os.path.getctime)
        try:
            with open(latest_log, "r", encoding="utf-8") as f:
                lines = f.readlines()
            log_text = "".join(lines[-40:])
            return {"success": True, "source": latest_log.name, "logs": log_text}
        except Exception as e:
            return {"success": False, "error": str(e)}
    return {"success": True, "source": "None", "logs": "No logs generated yet. Process a file to begin."}


# ─────────────────────── Configuration ───────────────────────
# Editable config keys with their types
EDITABLE_CONFIG = {
    "structure": {
        "bus_bars": ("BUS_BARS", int),
        "points_per_bar": ("POINTS_PER_BAR", int),
    },
    "thresholds": {
        "rule_a_threshold": ("RULE_A_THRESHOLD", float),
        "rule_a_percentage": ("RULE_A_PERCENTAGE", float),
        "rule_b_threshold": ("RULE_B_THRESHOLD", float),
        "max_rule_b_per_bar": ("MAX_RULE_B_PER_BAR", int),
        "rule_c_threshold": ("RULE_C_THRESHOLD", float),
        "max_rule_c_total": ("MAX_RULE_C_TOTAL", int),
        "max_rule_c_per_bar": ("MAX_RULE_C_PER_BAR", int),
    },
    "ocr": {
        "min_confidence": ("MIN_OCR_CONFIDENCE", float),
        "data_value_min": ("DATA_VALUE_MIN", float),
        "data_value_max": ("DATA_VALUE_MAX", float),
    },
    "verification": {
        "tolerance": ("VERIFY_TOLERANCE", float),
        "match_threshold": ("VERIFY_MATCH_THRESHOLD", float),
    },
    "intelligence": {
        "rag_enabled": ("ENABLE_RAG_CONTEXT", bool),
        "rag_top_k_similar": ("RAG_TOP_K_SIMILAR", int),
        "rag_similarity_threshold": ("RAG_SIMILARITY_THRESHOLD", float),
        "rag_llm_enabled": ("ENABLE_RAG_LLM_SUMMARY", bool),
        "llm_provider": ("LLM_PROVIDER", str),
        "ollama_model_name": ("OLLAMA_MODEL_NAME", str),
        "ollama_base_url": ("OLLAMA_BASE_URL", str),
        "llm_temperature": ("LLM_TEMPERATURE", float),
        "llm_max_tokens": ("LLM_MAX_TOKENS", int),
        "operator_feedback_enabled": ("ENABLE_OPERATOR_FEEDBACK", bool),
    },
}

@app.get("/api/config")
async def get_config():
    return {
        "structure": {
            "bus_bars": cfg_module.BUS_BARS,
            "points_per_bar": cfg_module.POINTS_PER_BAR,
            "total_points": cfg_module.TOTAL_POINTS,
        },
        "thresholds": {
            "rule_a_threshold": cfg_module.RULE_A_THRESHOLD,
            "rule_a_percentage": cfg_module.RULE_A_PERCENTAGE,
            "min_points_rule_a": cfg_module.MIN_POINTS_RULE_A,
            "rule_b_threshold": cfg_module.RULE_B_THRESHOLD,
            "max_rule_b_per_bar": cfg_module.MAX_RULE_B_PER_BAR,
            "rule_c_threshold": cfg_module.RULE_C_THRESHOLD,
            "max_rule_c_total": cfg_module.MAX_RULE_C_TOTAL,
            "max_rule_c_per_bar": cfg_module.MAX_RULE_C_PER_BAR,
        },
        "ocr": {
            "min_confidence": cfg_module.MIN_OCR_CONFIDENCE,
            "data_value_min": cfg_module.DATA_VALUE_MIN,
            "data_value_max": cfg_module.DATA_VALUE_MAX,
        },
        "verification": {
            "tolerance": cfg_module.VERIFY_TOLERANCE,
            "match_threshold": cfg_module.VERIFY_MATCH_THRESHOLD,
        },
        "intelligence": {
            "rag_enabled": cfg_module.ENABLE_RAG_CONTEXT,
            "rag_top_k_similar": cfg_module.RAG_TOP_K_SIMILAR,
            "rag_similarity_threshold": cfg_module.RAG_SIMILARITY_THRESHOLD,
            "rag_llm_enabled": cfg_module.ENABLE_RAG_LLM_SUMMARY,
            "llm_provider": cfg_module.LLM_PROVIDER,
            "ollama_model_name": cfg_module.OLLAMA_MODEL_NAME,
            "ollama_base_url": cfg_module.OLLAMA_BASE_URL,
            "llm_temperature": cfg_module.LLM_TEMPERATURE,
            "llm_max_tokens": cfg_module.LLM_MAX_TOKENS,
            "operator_feedback_enabled": cfg_module.ENABLE_OPERATOR_FEEDBACK,
        },
    }


@app.post("/api/feedback")
async def submit_feedback(request: Request):
    """Store operator feedback for a reviewed batch."""
    try:
        body = await request.json()
        batch_id = str(body.get("batch_id", "")).strip()
        if not batch_id:
            return JSONResponse(status_code=400, content={"success": False, "error": "batch_id is required"})

        feedback = {
            "root_cause": str(body.get("root_cause", "")).strip(),
            "operator_feedback": str(body.get("operator_feedback", "")).strip(),
            "feedback_confidence": str(body.get("feedback_confidence", "")).strip(),
            "reviewed_by": str(body.get("reviewed_by", "")).strip(),
            "reviewed_at": str(body.get("reviewed_at", datetime.now().isoformat(timespec="seconds"))).strip(),
            "action_taken": str(body.get("action_taken", "")).strip(),
        }

        updated = rag_history_db.update_feedback(batch_id, feedback)
        if not updated:
            return JSONResponse(status_code=404, content={"success": False, "error": "Batch not found"})

        return {"success": True, "batch_id": batch_id}
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "error": str(e)})


@app.post("/api/config")
async def update_config(request: Request):
    """Update configuration values at runtime. Changes persist in memory only."""
    try:
        body = await request.json()
        updated = []
        for section, fields in body.items():
            if section not in EDITABLE_CONFIG:
                continue
            for key, value in fields.items():
                if key not in EDITABLE_CONFIG[section]:
                    continue
                attr_name, type_fn = EDITABLE_CONFIG[section][key]
                new_val = type_fn(value)
                setattr(cfg_module, attr_name, new_val)
                updated.append(f"{attr_name} = {new_val}")

        # Recompute derived values
        cfg_module.TOTAL_POINTS = cfg_module.BUS_BARS * cfg_module.POINTS_PER_BAR
        cfg_module.MIN_POINTS_RULE_A = int(cfg_module.TOTAL_POINTS * cfg_module.RULE_A_PERCENTAGE)

        return {"success": True, "updated": updated}
    except Exception as e:
        return JSONResponse(status_code=500, content={"success": False, "error": str(e)})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("web_app:app", host="0.0.0.0", port=8000, reload=True)
