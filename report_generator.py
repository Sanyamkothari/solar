"""
Report Generator for the factory QC pipeline.
Outputs a standardized Excel report with highlighting for ease of operator reading.
Saves to the `/output` folder with the Batch ID.
"""
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from typing import List, Dict, Optional
from pathlib import Path
from config import (
    OUTPUT_DIR,
    RULE_A_THRESHOLD,
    RULE_B_THRESHOLD,
    RULE_C_THRESHOLD,
    ENABLE_RAG_CONTEXT,
)
from logger import logger

# Colors
RED_FILL = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
GREEN_FILL = PatternFill(start_color="99FF99", end_color="99FF99", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
LIGHT_BLUE_FILL = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")

class ReportGenerator:
    @staticmethod
    def generate_report(batch_id: str, matrix: List[List[float]], eval_report: Dict,
                        verification_report: Optional[Dict] = None, matrix_source: Optional[str] = None,
                        rag_context: Optional[Dict] = None) -> Path:
        """
        Generates the final Excel file.
        Includes sheets:
        1. Cleaned Data (with color highlights for failures).
        2. QC Summary (Pass/Fail metrics).
        3. Verification (if cross-verification was performed).
        4. RAG Context (if similar historical cases found).
        """
        if not matrix:
            raise ValueError(f"Cannot generate report for batch {batch_id}: matrix is empty.")

        wb = openpyxl.Workbook()
        
        # Sheet 1: Data Viewer
        ws_data = wb.active
        ws_data.title = "Cleaned Data"
        
        # Write Title & Decision
        ws_data["A1"] = f"Batch ID: {batch_id}"
        ws_data["A2"] = "Decision:"
        ws_data["B2"] = eval_report["decision"]
        
        # Color decision box
        if eval_report["decision"] == "APPROVED":
            ws_data["B2"].fill = GREEN_FILL
        elif eval_report["decision"] == "REJECTED":
            ws_data["B2"].fill = RED_FILL
        else:
            ws_data["B2"].fill = YELLOW_FILL
            
        ws_data["B2"].font = Font(bold=True)
        
        # Write Matrix with Rule Highlights
        start_row = 5
        ws_data.cell(row=start_row, column=1, value="s/n").font = Font(bold=True)
        for i in range(7):
            ws_data.cell(row=start_row, column=i+2, value=f"P{i+1}").font = Font(bold=True)
            
        for r_idx, row in enumerate(matrix, start=start_row+1):
            ws_data.cell(row=r_idx, column=1, value=f"Bus Bar {r_idx - start_row}")
            for c_idx, val in enumerate(row, start=2):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
                
                # Highlight failures
                if val <= RULE_C_THRESHOLD:
                    cell.fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid") # Darker red for severe fail
                elif val <= RULE_B_THRESHOLD:
                    cell.fill = RED_FILL
                elif val > RULE_A_THRESHOLD:
                    cell.fill = GREEN_FILL
        
        # Sheet 2: Metrics Summary
        ws_summary = wb.create_sheet(title="QC Summary")
        
        # Extract metrics
        metrics = eval_report.get("metrics", {})
        rule_a = metrics.get("rule_A", {})
        rule_b = metrics.get("rule_B", {})
        rule_c = metrics.get("rule_C", {})
        
        rows = [
            ["Metric", "Value", "Threshold/Rule", "Passed"],
            ["Rule A: >0.8 Points", rule_a.get("points_gt_08"), f">= {rule_a.get('required')}", str(rule_a.get("passed"))],
            ["Rule B: Max <=0.35 per bar", "See Data Sheet", "<= 2 per bar", str(rule_b.get("passed"))],
            ["Rule C: Total <=0.1 Points", rule_c.get("total_failures"), "<= 8 Total", str(rule_c.get("passed"))],
            ["Rule C: Max <=0.1 per bar", "See Data Sheet", "<= 1 per bar", ""]
        ]
        
        for r_idx, row_data in enumerate(rows, start=1):
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_summary.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = Font(bold=True)

        # Sheet 2 addendum: matrix source info
        if matrix_source:
            next_row = len(rows) + 3
            ws_summary.cell(row=next_row, column=1, value="Data Source:").font = Font(bold=True)
            ws_summary.cell(row=next_row, column=2, value=matrix_source)

        # Sheet 3: Verification Results (only if cross-verification was performed)
        if verification_report:
            ws_verify = wb.create_sheet(title="Verification")

            ws_verify["A1"] = "Cross-Verification: Image vs Excel"
            ws_verify["A1"].font = Font(bold=True, size=12)

            ws_verify["A3"] = "Result:"
            ws_verify["A3"].font = Font(bold=True)
            passed = verification_report["passed"]
            ws_verify["B3"] = "PASS" if passed else "FAIL"
            ws_verify["B3"].fill = GREEN_FILL if passed else RED_FILL
            ws_verify["B3"].font = Font(bold=True)

            info_rows = [
                ("Match Percentage", f"{verification_report['match_percentage']}%"),
                ("Matched Cells", f"{verification_report['matched_cells']} / {verification_report['total_cells']}"),
                ("Mismatches", str(verification_report["mismatch_count"])),
                ("Tolerance", f"±{verification_report['tolerance']}"),
            ]
            for i, (label, value) in enumerate(info_rows, start=5):
                ws_verify.cell(row=i, column=1, value=label).font = Font(bold=True)
                ws_verify.cell(row=i, column=2, value=value)

            # Mismatch detail table
            mismatches = verification_report.get("mismatches", [])
            if mismatches:
                header_row = 5 + len(info_rows) + 1
                headers = ["Bus Bar", "Point", "Image (OCR)", "Excel (Ref)", "Difference"]
                for c, h in enumerate(headers, start=1):
                    ws_verify.cell(row=header_row, column=c, value=h).font = Font(bold=True)

                for m_idx, m in enumerate(mismatches, start=header_row + 1):
                    ws_verify.cell(row=m_idx, column=1, value=m["bar"])
                    ws_verify.cell(row=m_idx, column=2, value=m["point"])
                    ws_verify.cell(row=m_idx, column=3, value=m["image_value"])
                    ws_verify.cell(row=m_idx, column=4, value=m["excel_value"])
                    diff_cell = ws_verify.cell(row=m_idx, column=5, value=m["difference"])
                    diff_cell.fill = RED_FILL

        # Sheet 4: RAG Context (similar historical cases)
        if ENABLE_RAG_CONTEXT and rag_context:
            ws_rag = wb.create_sheet(title="RAG Context")
            
            ws_rag["A1"] = "Similar Historical Cases (AI-Powered Context)"
            ws_rag["A1"].font = Font(bold=True, size=12)
            
            similar_batches = rag_context.get("similar_batches", [])
            if similar_batches:
                # Summary
                ws_rag["A3"] = "Current Decision:" 
                ws_rag["A3"].font = Font(bold=True)
                ws_rag["B3"] = eval_report.get("decision", "UNKNOWN")
                ws_rag["B3"].fill = LIGHT_BLUE_FILL
                
                # Similar cases table
                header_row = 5
                headers = ["Rank", "Batch ID", "Decision", "Similarity Score", "Shift"]
                for c, h in enumerate(headers, start=1):
                    ws_rag.cell(row=header_row, column=c, value=h).font = Font(bold=True)
                    ws_rag.cell(row=header_row, column=c).fill = LIGHT_BLUE_FILL
                
                for i, (batch_id_hist, similarity, metadata) in enumerate(similar_batches, start=1):
                    row = header_row + i
                    ws_rag.cell(row=row, column=1, value=i)
                    ws_rag.cell(row=row, column=2, value=batch_id_hist)
                    ws_rag.cell(row=row, column=3, value=metadata.get("decision", "UNKNOWN"))
                    sim_cell = ws_rag.cell(row=row, column=4, value=f"{similarity:.1%}")
                    sim_cell.number_format = '0%'
                    ws_rag.cell(row=row, column=5, value=metadata.get("shift", "UNKNOWN"))
                
                # Pattern analysis
                decision_pattern = rag_context.get("decision_pattern", {})
                if decision_pattern:
                    pattern_row = header_row + len(similar_batches) + 3
                    ws_rag.cell(row=pattern_row, column=1, value="Decision Pattern Analysis:").font = Font(bold=True)
                    ws_rag.cell(row=pattern_row + 1, column=1, value=f"Similar cases in history: {decision_pattern.get('total_cases', 0)}")
                    ws_rag.cell(row=pattern_row + 2, column=1, value=f"Frequency in last 100 batches: {decision_pattern.get('frequency', 'N/A')}")
            else:
                ws_rag["A3"] = "No similar historical cases found in database."
                ws_rag["A3"].font = Font(italic=True, color="808080")
                
        # Save output
        out_path = OUTPUT_DIR / f"{batch_id}_Report.xlsx"
        wb.save(out_path)
        
        logger.info(f"Report generated successfully: {out_path.name}")
        return out_path
