"""
Streamlit Dashboard for Factory QC Automation.
Multi-tab operator interface with upload, history, analytics, and live monitoring.
"""
import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import openpyxl
import os
import tempfile
from pathlib import Path
from datetime import datetime

# Adjust Python path
import sys
BASE_DIR = Path(__file__).parent
if str(BASE_DIR) not in sys.path:
    sys.path.append(str(BASE_DIR))

from config import (
    INPUT_DIR, PROCESSED_DIR, FAILED_DIR, OUTPUT_DIR, LOGS_DIR,
    RULE_A_THRESHOLD, RULE_B_THRESHOLD, RULE_C_THRESHOLD,
    MIN_POINTS_RULE_A, MAX_RULE_B_PER_BAR, MAX_RULE_C_TOTAL, MAX_RULE_C_PER_BAR,
    VERIFY_TOLERANCE, BUS_BARS, POINTS_PER_BAR, TOTAL_POINTS,
    RULE_A_PERCENTAGE, VERIFY_MATCH_THRESHOLD, MIN_BUS_BARS,
)
from input_handler import InputHandler
from main import process_file

st.set_page_config(page_title="Factory QC Dashboard", layout="wide", page_icon="🏭")

# ───────────────────────────── CSS ─────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

    /* Metric cards */
    .metric-card {
        background: linear-gradient(135deg, #1a1a2e 0%, #16213e 100%);
        padding: 20px 16px; border-radius: 14px;
        border: 1px solid rgba(255,255,255,0.06);
        text-align: center; margin-bottom: 6px;
    }
    .metric-card h3 { color: #a0aec0; font-size: 13px; margin-bottom: 2px; font-weight: 600; letter-spacing: .3px; }
    .metric-value-green { color: #48BB78; font-size: 36px; font-weight: 700; }
    .metric-value-red   { color: #FC8181; font-size: 36px; font-weight: 700; }
    .metric-value-blue  { color: #63B3ED; font-size: 36px; font-weight: 700; }
    .metric-value-amber { color: #F6E05E; font-size: 36px; font-weight: 700; }

    /* Pipeline step rows */
    .step-row {
        display: flex; align-items: center;
        padding: 10px 16px; margin: 4px 0;
        border-radius: 10px; font-size: 15px;
        border-left: 4px solid transparent;
    }
    .step-pass { background: rgba(72,187,120,0.1); border-left-color: #48BB78; }
    .step-fail { background: rgba(252,129,129,0.1); border-left-color: #FC8181; }
    .step-warn { background: rgba(246,224,94,0.1); border-left-color: #F6E05E; }
    .step-info { background: rgba(99,179,237,0.1); border-left-color: #63B3ED; }
    .step-name   { font-weight: 600; color: #E2E8F0; min-width: 180px; }
    .step-status { font-weight: 700; min-width: 120px; text-align: center; }
    .step-detail { color: #A0AEC0; font-size: 13px; flex: 1; }
    .step-time   { color: #718096; font-size: 12px; min-width: 70px; text-align: right; }

    /* Decision banners */
    .decision-banner {
        text-align: center; padding: 20px; border-radius: 16px;
        font-size: 28px; font-weight: 700; margin: 16px 0; letter-spacing: 2px;
    }
    .decision-approved { background: linear-gradient(135deg, #22543d, #276749); color: #C6F6D5; }
    .decision-rejected { background: linear-gradient(135deg, #742a2a, #9b2c2c); color: #FED7D7; }
    .decision-error    { background: linear-gradient(135deg, #744210, #975a16); color: #FEFCBF; }
    .decision-review   { background: linear-gradient(135deg, #2a4365, #2b6cb0); color: #BEE3F8; }

    /* History table row badges */
    .badge-approved { background:#22543d; color:#C6F6D5; padding:3px 10px; border-radius:6px; font-weight:600; font-size:13px; }
    .badge-rejected { background:#742a2a; color:#FED7D7; padding:3px 10px; border-radius:6px; font-weight:600; font-size:13px; }
    .badge-error    { background:#744210; color:#FEFCBF; padding:3px 10px; border-radius:6px; font-weight:600; font-size:13px; }
    .badge-review   { background:#2a4365; color:#BEE3F8; padding:3px 10px; border-radius:6px; font-weight:600; font-size:13px; }

    /* Activity feed */
    .activity-item {
        padding: 10px 14px; margin: 4px 0; border-radius: 8px;
        background: rgba(255,255,255,0.03); border-left: 3px solid #63B3ED;
        font-size: 14px;
    }
    .activity-item .ts { color: #718096; font-size: 12px; }
    .activity-item .fname { color: #E2E8F0; font-weight: 500; }

    /* Config table */
    .config-row {
        display: flex; justify-content: space-between; align-items: center;
        padding: 10px 16px; margin: 3px 0; border-radius: 8px;
        background: rgba(255,255,255,0.03); font-size: 14px;
    }
    .config-label { color: #A0AEC0; font-weight: 500; }
    .config-value { color: #E2E8F0; font-weight: 700; font-family: monospace; font-size: 15px; }
</style>
""", unsafe_allow_html=True)


# ───────────────────────────── HELPERS ─────────────────────────────

@st.cache_data(ttl=30)
def get_counts():
    """Return file counts for KPI cards."""
    return {
        "processed": len(list(PROCESSED_DIR.glob('*'))),
        "reports": len(list(OUTPUT_DIR.glob('*.xlsx'))),
        "pending": len(InputHandler.get_pending_files()),
    }


@st.cache_data(ttl=60)
def load_report_history():
    """Read all output reports and return summary dataframe."""
    reports = sorted(OUTPUT_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    rows = []
    for rp in reports:
        try:
            wb = openpyxl.load_workbook(rp, read_only=True, data_only=True)
            ws_data = wb["Cleaned Data"]
            ws_summary = wb["QC Summary"]

            batch_id = (ws_data["A1"].value or "").replace("Batch ID: ", "")
            decision = ws_data["B2"].value or "UNKNOWN"

            rule_a_pts = ws_summary.cell(row=2, column=2).value
            rule_a_pass = ws_summary.cell(row=2, column=4).value
            rule_c_total = ws_summary.cell(row=4, column=2).value

            # Read matrix for stats
            matrix_vals = []
            for row in ws_data.iter_rows(min_row=6, min_col=2, max_col=8, values_only=True):
                vals = [v for v in row if isinstance(v, (int, float))]
                matrix_vals.extend(vals)

            # Get data source
            source = ""
            for row in ws_summary.iter_rows(min_row=6, max_row=15, max_col=2, values_only=True):
                if row[0] == "Data Source:":
                    source = row[1] or ""
                    break

            # Parse timestamp from batch id
            ts = None
            parts = batch_id.split("_")
            if len(parts) >= 3:
                try:
                    ts = datetime.strptime(f"{parts[1]}_{parts[2]}", "%Y%m%d_%H%M%S")
                except ValueError:
                    pass

            wb.close()
            rows.append({
                "batch_id": batch_id,
                "decision": decision,
                "timestamp": ts,
                "rule_a_points": rule_a_pts,
                "rule_a_passed": str(rule_a_pass) == "True",
                "rule_c_total": rule_c_total,
                "total_points": len(matrix_vals),
                "avg_force": round(sum(matrix_vals) / len(matrix_vals), 3) if matrix_vals else 0,
                "min_force": round(min(matrix_vals), 3) if matrix_vals else 0,
                "max_force": round(max(matrix_vals), 3) if matrix_vals else 0,
                "source": source,
                "file": rp.name,
                "path": str(rp),
            })
        except Exception:
            continue
    return pd.DataFrame(rows) if rows else pd.DataFrame()


def render_metric_card(label, value, color_class):
    st.markdown(
        f"<div class='metric-card'><h3>{label}</h3>"
        f"<span class='{color_class}'>{value}</span></div>",
        unsafe_allow_html=True,
    )


def render_decision_banner(decision):
    banner_map = {
        "APPROVED": ("decision-approved", "✅"),
        "REJECTED": ("decision-rejected", "❌"),
        "DATA_ERROR": ("decision-error", "⚠️"),
        "VERIFICATION_FAILED": ("decision-error", "🔀"),
        "MANUAL_REVIEW_REQUIRED": ("decision-review", "🔍"),
    }
    css, icon = banner_map.get(decision, ("decision-review", "🔍"))
    st.markdown(
        f"<div class='decision-banner {css}'>{icon}  {decision}</div>",
        unsafe_allow_html=True,
    )


def decision_badge(decision):
    badge_map = {
        "APPROVED": "badge-approved",
        "REJECTED": "badge-rejected",
        "DATA_ERROR": "badge-error",
        "VERIFICATION_FAILED": "badge-error",
        "MANUAL_REVIEW_REQUIRED": "badge-review",
    }
    css = badge_map.get(decision, "badge-review")
    return f"<span class='{css}'>{decision}</span>"


# ───────────────────────────── SIDEBAR NAV ─────────────────────────────

with st.sidebar:
    st.markdown("## 🏭 Factory QC")
    st.caption("Solder Point Peel Test Automation")
    st.markdown("---")
    page = st.radio(
        "Navigation",
        ["🏠 Overview", "📂 Run Inspection", "📋 Report History", "📊 Analytics", "⚙️ Settings", "📜 Logs"],
        label_visibility="collapsed",
    )
    st.markdown("---")
    counts = get_counts()
    st.metric("Reports Generated", counts["reports"])
    st.metric("Processed Files", counts["processed"])
    st.markdown("---")
    if st.button("🔄 Refresh", use_container_width=True):
        st.cache_data.clear()
        st.rerun()


# ╔══════════════════════════════════════════════════════════════════╗
# ║                       PAGE: OVERVIEW                            ║
# ╚══════════════════════════════════════════════════════════════════╝
if page == "🏠 Overview":
    st.title("🏠 Dashboard Overview")
    st.caption(f"Manufacturing QC Pipeline — {datetime.now().strftime('%A, %B %d, %Y  %H:%M')}")

    # KPI row
    c1, c2, c3 = st.columns(3)
    with c1: render_metric_card("✅ Processed", counts["processed"], "metric-value-green")
    with c2: render_metric_card("📊 Reports", counts["reports"], "metric-value-blue")
    with c3: render_metric_card("⏳ Pending", counts["pending"], "metric-value-amber")

    st.markdown("---")

    # Two-column layout: recent activity + quick stats
    col_left, col_right = st.columns([3, 2])

    with col_left:
        st.subheader("📈 Recent Batch Results")
        df_hist = load_report_history()
        if not df_hist.empty:
            recent = df_hist.head(10)
            for _, row in recent.iterrows():
                ts_str = row["timestamp"].strftime("%b %d, %H:%M") if pd.notna(row["timestamp"]) else "—"
                badge = decision_badge(row["decision"])
                avg = row["avg_force"]
                st.markdown(
                    f"<div class='activity-item'>"
                    f"<span class='ts'>{ts_str}</span> &nbsp; {badge} &nbsp; "
                    f"<span class='fname'>{row['batch_id']}</span> &nbsp; "
                    f"<span style='color:#A0AEC0; font-size:13px;'>Avg: {avg}N · {row['total_points']} pts</span>"
                    f"</div>",
                    unsafe_allow_html=True,
                )
        else:
            st.info("No reports generated yet. Upload a file to start.")

    with col_right:
        st.subheader("📊 Pass/Fail Breakdown")
        if not df_hist.empty:
            decision_counts = df_hist["decision"].value_counts()
            colors = {"APPROVED": "#48BB78", "REJECTED": "#FC8181", "DATA_ERROR": "#F6E05E",
                      "MANUAL_REVIEW_REQUIRED": "#63B3ED", "VERIFICATION_FAILED": "#ED8936"}
            fig = px.pie(
                names=decision_counts.index,
                values=decision_counts.values,
                color=decision_counts.index,
                color_discrete_map=colors,
                hole=0.45,
            )
            fig.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#E2E8F0"), showlegend=True,
                legend=dict(orientation="h", yanchor="bottom", y=-0.2),
                margin=dict(t=20, b=20, l=20, r=20), height=300,
            )
            fig.update_traces(textinfo="value+percent", textfont_size=13)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("No data yet.")

        # Quick force distribution
        if not df_hist.empty:
            st.subheader("⚡ Avg Force Trend")
            trend_df = df_hist.dropna(subset=["timestamp"]).sort_values("timestamp")
            if not trend_df.empty:
                fig2 = px.line(
                    trend_df, x="timestamp", y="avg_force",
                    markers=True, color_discrete_sequence=["#63B3ED"],
                )
                fig2.add_hline(y=RULE_A_THRESHOLD, line_dash="dash", line_color="#48BB78",
                               annotation_text=f"Rule A ({RULE_A_THRESHOLD}N)")
                fig2.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#E2E8F0"), showlegend=False,
                    xaxis_title="", yaxis_title="Avg Force (N)",
                    margin=dict(t=20, b=20, l=20, r=20), height=250,
                )
                st.plotly_chart(fig2, use_container_width=True)


# ╔══════════════════════════════════════════════════════════════════╗
# ║                    PAGE: RUN INSPECTION                         ║
# ╚══════════════════════════════════════════════════════════════════╝
elif page == "📂 Run Inspection":
    st.title("📂 Run Batch Inspection")
    st.caption("Upload an image or Excel file to run the full QC pipeline")

    # Upload section
    col_up1, col_up2 = st.columns(2)
    with col_up1:
        st.markdown("##### 📷 Test File (Image or Excel)")
        uploaded_file = st.file_uploader(
            "Drop an Image (PNG/JPG) or Excel (XLSX/XLS)",
            type=["png", "jpg", "jpeg", "xlsx", "xls"],
            key="main_upload",
        )
    with col_up2:
        st.markdown("##### 📎 Reference Excel *(optional)*")
        st.caption("Upload matching Excel to cross-verify OCR accuracy")
        excel_ref_file = st.file_uploader(
            "Reference Excel (XLSX/XLS)",
            type=["xlsx", "xls"],
            key="excel_ref",
        )

    if uploaded_file is not None:
        # Preview uploaded file info
        st.markdown(
            f"**File:** {uploaded_file.name}  |  **Size:** {uploaded_file.size / 1024:.1f} KB  |  "
            f"**Type:** {uploaded_file.type}"
        )

        if st.button("🚀  Start QC Pipeline", use_container_width=True, type="primary"):
            tmp_dir = Path(tempfile.mkdtemp(prefix="qc_upload_"))
            file_path = tmp_dir / uploaded_file.name
            with open(file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())

            excel_ref_path = None
            if excel_ref_file is not None:
                excel_ref_path = tmp_dir / f"ref_{excel_ref_file.name}"
                with open(excel_ref_path, "wb") as f:
                    f.write(excel_ref_file.getbuffer())

            with st.spinner("Running factory QC pipeline…"):
                try:
                    result = process_file(file_path, excel_ref_path=excel_ref_path)
                except Exception as e:
                    st.error(f"Critical pipeline failure: {e}")
                    result = None

            if result:
                # Clear cache so new report shows up immediately
                st.cache_data.clear()

                # Decision banner
                render_decision_banner(result.get("decision", "UNKNOWN"))
                st.caption(f"Batch ID: `{result.get('batch_id', 'N/A')}`")

                # Matrix source
                matrix_source = result.get("matrix_source")
                if matrix_source:
                    st.info(f"📋 **Data Source:** {matrix_source}")

                # Cross-verification
                vr = result.get("verification_report")
                if vr:
                    st.subheader("🔀 Cross-Verification: Image vs Excel")
                    vc1, vc2, vc3 = st.columns(3)
                    with vc1:
                        color = "metric-value-green" if vr["passed"] else "metric-value-red"
                        render_metric_card("Match Rate", f"{vr['match_percentage']}%", color)
                    with vc2:
                        render_metric_card("Matched Cells", f"{vr['matched_cells']}/{vr['total_cells']}", "metric-value-blue")
                    with vc3:
                        mc = "metric-value-green" if vr["mismatch_count"] == 0 else "metric-value-red"
                        render_metric_card("Mismatches", vr["mismatch_count"], mc)

                    if vr["passed"]:
                        st.success(f"✅ Verification PASSED — within ±{vr['tolerance']} tolerance.")
                    else:
                        st.error(f"❌ Verification FAILED — {vr['mismatch_count']} cells differ beyond ±{vr['tolerance']}. Excel used as ground truth.")

                    if vr["mismatches"]:
                        with st.expander("View mismatch details"):
                            mismatch_df = pd.DataFrame(vr["mismatches"])
                            mismatch_df.columns = ["Bus Bar", "Point", "Image (OCR)", "Excel (Ref)", "Difference"]
                            st.dataframe(mismatch_df, use_container_width=True, hide_index=True)

                # Pipeline steps & Data matrix in two tabs
                tab_steps, tab_matrix, tab_rules = st.tabs(["🔗 Pipeline Steps", "🔬 Data Matrix", "📋 Rule Summary"])

                with tab_steps:
                    for step in result.get("steps", []):
                        status = step["status"]
                        if "PASS" in status or "DONE" in status:
                            row_class = "step-pass"
                        elif "FAIL" in status:
                            row_class = "step-fail"
                        elif "MOVED" in status or "OVERRIDE" in status or "PARTIAL" in status:
                            row_class = "step-warn"
                        else:
                            row_class = "step-info"
                        st.markdown(
                            f"<div class='step-row {row_class}'>"
                            f"<span class='step-name'>{step['name']}</span>"
                            f"<span class='step-status'>{step['status']}</span>"
                            f"<span class='step-detail'>{step['detail']}</span>"
                            f"<span class='step-time'>{step['time']}</span>"
                            f"</div>",
                            unsafe_allow_html=True,
                        )

                with tab_matrix:
                    matrix = result.get("matrix")
                    if matrix:
                        rows_m = len(matrix)
                        cols_m = len(matrix[0]) if matrix else 0
                        st.caption(f"{rows_m} bus bars × {cols_m} measurement points")

                        df = pd.DataFrame(
                            matrix,
                            columns=[f"P{i+1}" for i in range(cols_m)],
                            index=[f"Bar {i+1}" for i in range(rows_m)],
                        )

                        def color_cells(val):
                            if val <= RULE_C_THRESHOLD:
                                return "background-color: #742a2a; color: #FED7D7; font-weight: bold;"
                            elif val <= RULE_B_THRESHOLD:
                                return "background-color: #9b2c2c; color: #FED7D7;"
                            elif val > RULE_A_THRESHOLD:
                                return "background-color: #22543d; color: #C6F6D5;"
                            else:
                                return "background-color: #744210; color: #FEFCBF;"

                        styled = df.style.map(color_cells).format("{:.3f}")
                        st.dataframe(styled, use_container_width=True, height=620)

                        # Mini force distribution chart
                        all_vals = [v for row in matrix for v in row]
                        fig_dist = px.histogram(
                            x=all_vals, nbins=30,
                            labels={"x": "Force (N)", "y": "Count"},
                            color_discrete_sequence=["#63B3ED"],
                        )
                        fig_dist.add_vline(x=RULE_A_THRESHOLD, line_dash="dash", line_color="#48BB78",
                                           annotation_text="0.8N")
                        fig_dist.add_vline(x=RULE_B_THRESHOLD, line_dash="dash", line_color="#F6E05E",
                                           annotation_text="0.35N")
                        fig_dist.add_vline(x=RULE_C_THRESHOLD, line_dash="dash", line_color="#FC8181",
                                           annotation_text="0.1N")
                        fig_dist.update_layout(
                            paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                            font=dict(color="#E2E8F0"), showlegend=False,
                            xaxis_title="Force (N)", yaxis_title="Count",
                            margin=dict(t=30, b=30, l=30, r=30), height=280,
                            title="Force Distribution",
                        )
                        st.plotly_chart(fig_dist, use_container_width=True)
                    else:
                        st.warning("No matrix data available.")

                with tab_rules:
                    eval_report = result.get("eval_report")
                    if eval_report:
                        metrics = eval_report["metrics"]
                        # Visual rule cards
                        rc1, rc2, rc3 = st.columns(3)
                        with rc1:
                            passed_a = metrics["rule_A"]["passed"]
                            icon_a = "✅" if passed_a else "❌"
                            pts = metrics["rule_A"]["points_gt_08"]
                            req = metrics["rule_A"]["required"]
                            st.markdown(f"### {icon_a} Rule A")
                            st.markdown(f"**{pts} / {req}** points > 0.8N")
                            st.progress(min(pts / max(req, 1), 1.0))
                        with rc2:
                            passed_b = metrics["rule_B"]["passed"]
                            icon_b = "✅" if passed_b else "❌"
                            max_b = max(metrics["rule_B"]["failures_per_bar"].values()) if metrics["rule_B"]["failures_per_bar"] else 0
                            st.markdown(f"### {icon_b} Rule B")
                            st.markdown(f"Max ≤0.35N per bar: **{max_b}** (limit: {MAX_RULE_B_PER_BAR})")
                            st.progress(1.0 - min(max_b / max(MAX_RULE_B_PER_BAR + 1, 1), 1.0))
                        with rc3:
                            passed_c = metrics["rule_C"]["passed"]
                            icon_c = "✅" if passed_c else "❌"
                            total_c = metrics["rule_C"]["total_failures"]
                            st.markdown(f"### {icon_c} Rule C")
                            st.markdown(f"Total ≤0.1N: **{total_c}** (limit: {MAX_RULE_C_TOTAL})")
                            st.progress(1.0 - min(total_c / max(MAX_RULE_C_TOTAL + 1, 1), 1.0))

                # Download report
                report_path = result.get("report_path")
                if report_path and os.path.exists(report_path):
                    st.markdown("---")
                    with open(report_path, "rb") as f:
                        st.download_button(
                            "📥  Download Excel Report",
                            data=f.read(),
                            file_name=os.path.basename(report_path),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
    else:
        # Empty state
        st.markdown("---")
        st.markdown(
            "<div style='text-align:center; padding:60px 20px; color:#718096;'>"
            "<p style='font-size:48px;'>📷</p>"
            "<p style='font-size:18px; font-weight:600;'>Upload an image or Excel file to begin inspection</p>"
            "<p style='font-size:14px;'>Supports PNG, JPG, XLSX, XLS — max 50 MB</p>"
            "</div>",
            unsafe_allow_html=True,
        )


# ╔══════════════════════════════════════════════════════════════════╗
# ║                    PAGE: REPORT HISTORY                         ║
# ╚══════════════════════════════════════════════════════════════════╝
elif page == "📋 Report History":
    st.title("📋 Report History")
    st.caption("Browse, search, and download past QC reports")

    df_hist = load_report_history()

    if df_hist.empty:
        st.info("No reports found. Process files to generate reports.")
    else:
        # Filters
        col_f1, col_f2, col_f3 = st.columns(3)
        with col_f1:
            decision_filter = st.multiselect(
                "Filter by Decision",
                options=df_hist["decision"].unique().tolist(),
                default=df_hist["decision"].unique().tolist(),
            )
        with col_f2:
            source_filter = st.multiselect(
                "Filter by Source",
                options=[s for s in df_hist["source"].unique().tolist() if s],
                default=[s for s in df_hist["source"].unique().tolist() if s],
            )
        with col_f3:
            search_term = st.text_input("🔍 Search Batch ID", "")

        filtered = df_hist[df_hist["decision"].isin(decision_filter)]
        if source_filter:
            filtered = filtered[filtered["source"].isin(source_filter)]
        if search_term:
            filtered = filtered[filtered["batch_id"].str.contains(search_term, case=False, na=False)]

        st.markdown(f"**Showing {len(filtered)} of {len(df_hist)} reports**")

        # Display table
        for _, row in filtered.iterrows():
            ts_str = row["timestamp"].strftime("%Y-%m-%d  %H:%M:%S") if pd.notna(row["timestamp"]) else "—"
            badge = decision_badge(row["decision"])

            with st.expander(f"{ts_str}  —  {row['batch_id']}  |  {row['decision']}"):
                mc1, mc2, mc3, mc4 = st.columns(4)
                with mc1:
                    st.markdown(f"**Decision**")
                    st.markdown(badge, unsafe_allow_html=True)
                with mc2:
                    st.metric("Avg Force", f"{row['avg_force']}N")
                with mc3:
                    st.metric("Min / Max", f"{row['min_force']} / {row['max_force']}N")
                with mc4:
                    st.metric("Rule A Points", f"{row['rule_a_points']}")

                st.caption(f"Source: {row['source']}  |  Points: {row['total_points']}  |  Rule C failures: {row['rule_c_total']}")

                # Download button for this specific report
                if os.path.exists(row["path"]):
                    with open(row["path"], "rb") as f:
                        st.download_button(
                            f"📥 Download {row['file']}",
                            data=f.read(),
                            file_name=row["file"],
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"dl_{row['batch_id']}",
                        )

                # Show matrix from report
                if st.checkbox("View data matrix", key=f"matrix_{row['batch_id']}"):
                    try:
                        wb = openpyxl.load_workbook(row["path"], read_only=True, data_only=True)
                        ws = wb["Cleaned Data"]
                        matrix_rows = []
                        for r in ws.iter_rows(min_row=6, min_col=2, max_col=8, values_only=True):
                            vals = [v if isinstance(v, (int, float)) else 0 for v in r]
                            if any(v != 0 for v in vals):
                                matrix_rows.append(vals)
                        wb.close()
                        if matrix_rows:
                            df_m = pd.DataFrame(
                                matrix_rows,
                                columns=[f"P{i+1}" for i in range(len(matrix_rows[0]))],
                                index=[f"Bar {i+1}" for i in range(len(matrix_rows))],
                            )

                            def color_cells_hist(val):
                                if val <= RULE_C_THRESHOLD:
                                    return "background-color: #742a2a; color: #FED7D7; font-weight: bold;"
                                elif val <= RULE_B_THRESHOLD:
                                    return "background-color: #9b2c2c; color: #FED7D7;"
                                elif val > RULE_A_THRESHOLD:
                                    return "background-color: #22543d; color: #C6F6D5;"
                                else:
                                    return "background-color: #744210; color: #FEFCBF;"

                            styled_m = df_m.style.map(color_cells_hist).format("{:.3f}")
                            st.dataframe(styled_m, use_container_width=True)
                    except Exception as e:
                        st.error(f"Could not read matrix: {e}")


# ╔══════════════════════════════════════════════════════════════════╗
# ║                      PAGE: ANALYTICS                            ║
# ╚══════════════════════════════════════════════════════════════════╝
elif page == "📊 Analytics":
    st.title("📊 Analytics & Trends")
    st.caption("Quality trends and performance insights across all batches")

    df_hist = load_report_history()

    if df_hist.empty:
        st.info("No data available yet. Process files to see analytics.")
    else:
        df_sorted = df_hist.dropna(subset=["timestamp"]).sort_values("timestamp")

        # Top summary row
        total_batches = len(df_hist)
        approved = len(df_hist[df_hist["decision"] == "APPROVED"])
        rejected = len(df_hist[df_hist["decision"] == "REJECTED"])
        approval_rate = round(approved / max(total_batches, 1) * 100, 1)

        s1, s2, s3, s4 = st.columns(4)
        with s1: render_metric_card("Total Batches", total_batches, "metric-value-blue")
        with s2: render_metric_card("Approved", approved, "metric-value-green")
        with s3: render_metric_card("Rejected", rejected, "metric-value-red")
        with s4: render_metric_card("Approval Rate", f"{approval_rate}%", "metric-value-green" if approval_rate >= 80 else "metric-value-amber")

        st.markdown("---")

        # Charts in two columns
        chart1, chart2 = st.columns(2)

        with chart1:
            st.subheader("📈 Average Force Over Time")
            if not df_sorted.empty:
                fig_force = go.Figure()
                fig_force.add_trace(go.Scatter(
                    x=df_sorted["timestamp"], y=df_sorted["avg_force"],
                    mode="lines+markers", name="Avg Force",
                    line=dict(color="#63B3ED", width=2),
                    marker=dict(size=7),
                ))
                fig_force.add_trace(go.Scatter(
                    x=df_sorted["timestamp"], y=df_sorted["min_force"],
                    mode="lines", name="Min Force",
                    line=dict(color="#FC8181", width=1, dash="dot"),
                ))
                fig_force.add_trace(go.Scatter(
                    x=df_sorted["timestamp"], y=df_sorted["max_force"],
                    mode="lines", name="Max Force",
                    line=dict(color="#48BB78", width=1, dash="dot"),
                ))
                fig_force.add_hline(y=RULE_A_THRESHOLD, line_dash="dash", line_color="rgba(72,187,120,0.5)",
                                    annotation_text="Rule A (0.8N)")
                fig_force.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#E2E8F0"),
                    xaxis_title="", yaxis_title="Force (N)",
                    margin=dict(t=20, b=20, l=40, r=20), height=350,
                    legend=dict(orientation="h", yanchor="bottom", y=1.02),
                )
                st.plotly_chart(fig_force, use_container_width=True)

        with chart2:
            st.subheader("📊 Decision Distribution")
            decision_counts = df_hist["decision"].value_counts()
            colors_map = {"APPROVED": "#48BB78", "REJECTED": "#FC8181", "DATA_ERROR": "#F6E05E",
                          "MANUAL_REVIEW_REQUIRED": "#63B3ED", "VERIFICATION_FAILED": "#ED8936"}
            fig_bar = px.bar(
                x=decision_counts.index, y=decision_counts.values,
                color=decision_counts.index, color_discrete_map=colors_map,
                labels={"x": "Decision", "y": "Count"},
            )
            fig_bar.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#E2E8F0"), showlegend=False,
                margin=dict(t=20, b=20, l=40, r=20), height=350,
            )
            st.plotly_chart(fig_bar, use_container_width=True)

        st.markdown("---")

        chart3, chart4 = st.columns(2)

        with chart3:
            st.subheader("🎯 Rule A Compliance")
            if not df_sorted.empty and "rule_a_points" in df_sorted.columns:
                fig_ra = go.Figure()
                fig_ra.add_trace(go.Bar(
                    x=df_sorted["batch_id"].str[-8:],
                    y=df_sorted["rule_a_points"],
                    marker_color=["#48BB78" if p else "#FC8181" for p in df_sorted["rule_a_passed"]],
                    name="Points > 0.8N",
                ))
                fig_ra.add_hline(y=MIN_POINTS_RULE_A, line_dash="dash", line_color="#F6E05E",
                                 annotation_text=f"Required ({MIN_POINTS_RULE_A})")
                fig_ra.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#E2E8F0"), showlegend=False,
                    xaxis_title="Batch (last 8 chars)", yaxis_title="Points > 0.8N",
                    margin=dict(t=20, b=20, l=40, r=20), height=350,
                )
                st.plotly_chart(fig_ra, use_container_width=True)

        with chart4:
            st.subheader("⚠️ Rule C Failures")
            if not df_sorted.empty and "rule_c_total" in df_sorted.columns:
                fig_rc = go.Figure()
                fig_rc.add_trace(go.Bar(
                    x=df_sorted["batch_id"].str[-8:],
                    y=df_sorted["rule_c_total"],
                    marker_color=["#48BB78" if v <= MAX_RULE_C_TOTAL else "#FC8181"
                                  for v in df_sorted["rule_c_total"].fillna(0)],
                    name="Points ≤ 0.1N",
                ))
                fig_rc.add_hline(y=MAX_RULE_C_TOTAL, line_dash="dash", line_color="#FC8181",
                                 annotation_text=f"Limit ({MAX_RULE_C_TOTAL})")
                fig_rc.update_layout(
                    paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="#E2E8F0"), showlegend=False,
                    xaxis_title="Batch (last 8 chars)", yaxis_title="Total ≤ 0.1N failures",
                    margin=dict(t=20, b=20, l=40, r=20), height=350,
                )
                st.plotly_chart(fig_rc, use_container_width=True)

        # Force distribution across all batches
        st.markdown("---")
        st.subheader("🔬 Global Force Distribution (All Batches)")
        all_forces = []
        for rp in sorted(OUTPUT_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)[:20]:
            try:
                wb = openpyxl.load_workbook(rp, read_only=True, data_only=True)
                ws = wb["Cleaned Data"]
                for row in ws.iter_rows(min_row=6, min_col=2, max_col=8, values_only=True):
                    for v in row:
                        if isinstance(v, (int, float)):
                            all_forces.append(v)
                wb.close()
            except Exception:
                continue

        if all_forces:
            fig_global = px.histogram(
                x=all_forces, nbins=50,
                labels={"x": "Force (N)", "y": "Count"},
                color_discrete_sequence=["#63B3ED"],
            )
            fig_global.add_vline(x=RULE_A_THRESHOLD, line_dash="dash", line_color="#48BB78",
                                 annotation_text="0.8N (Rule A)")
            fig_global.add_vline(x=RULE_B_THRESHOLD, line_dash="dash", line_color="#F6E05E",
                                 annotation_text="0.35N (Rule B)")
            fig_global.add_vline(x=RULE_C_THRESHOLD, line_dash="dash", line_color="#FC8181",
                                 annotation_text="0.1N (Rule C)")
            fig_global.update_layout(
                paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
                font=dict(color="#E2E8F0"), showlegend=False,
                xaxis_title="Peel Force (N)", yaxis_title="Frequency",
                margin=dict(t=30, b=30, l=40, r=20), height=350,
                title=f"Distribution of {len(all_forces)} measurements across last 20 batches",
            )
            st.plotly_chart(fig_global, use_container_width=True)


# ╔══════════════════════════════════════════════════════════════════╗
# ║                       PAGE: SETTINGS                            ║
# ╚══════════════════════════════════════════════════════════════════╝
elif page == "⚙️ Settings":
    st.title("⚙️ System Configuration")
    st.caption("Current QC thresholds and system parameters (read-only)")

    col_s1, col_s2 = st.columns(2)

    with col_s1:
        st.subheader("📏 Quality Rules")
        config_items = [
            ("Rule A — Min Force Threshold", f"{RULE_A_THRESHOLD} N"),
            ("Rule A — Required % of Points", f"{RULE_A_PERCENTAGE * 100}%"),
            ("Rule A — Min Points Required", str(MIN_POINTS_RULE_A)),
            ("Rule B — Low-Force Threshold", f"{RULE_B_THRESHOLD} N"),
            ("Rule B — Max Failures Per Bar", str(MAX_RULE_B_PER_BAR)),
            ("Rule C — Critical Threshold", f"{RULE_C_THRESHOLD} N"),
            ("Rule C — Max Total Failures", str(MAX_RULE_C_TOTAL)),
            ("Rule C — Max Failures Per Bar", str(MAX_RULE_C_PER_BAR)),
        ]
        for label, value in config_items:
            st.markdown(
                f"<div class='config-row'>"
                f"<span class='config-label'>{label}</span>"
                f"<span class='config-value'>{value}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )

    with col_s2:
        st.subheader("🔧 System Parameters")
        sys_items = [
            ("Expected Bus Bars", str(BUS_BARS)),
            ("Points Per Bar", str(POINTS_PER_BAR)),
            ("Total Expected Points", str(TOTAL_POINTS)),
            ("Min Bus Bars (Partial)", str(MIN_BUS_BARS)),
            ("Cross-Verify Tolerance", f"±{VERIFY_TOLERANCE}"),
            ("Cross-Verify Match Threshold", f"{VERIFY_MATCH_THRESHOLD * 100}%"),
            ("Max File Size", "50 MB"),
        ]
        for label, value in sys_items:
            st.markdown(
                f"<div class='config-row'>"
                f"<span class='config-label'>{label}</span>"
                f"<span class='config-value'>{value}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )

        st.subheader("📁 Directories")
        dir_items = [
            ("Input", str(INPUT_DIR)),
            ("Processed", str(PROCESSED_DIR)),
            ("Failed", str(FAILED_DIR)),
            ("Output", str(OUTPUT_DIR)),
            ("Logs", str(LOGS_DIR)),
        ]
        for label, value in dir_items:
            st.markdown(
                f"<div class='config-row'>"
                f"<span class='config-label'>{label}</span>"
                f"<span class='config-value' style='font-size:12px;'>{value}</span>"
                f"</div>",
                unsafe_allow_html=True,
            )

    st.markdown("---")
    st.subheader("🎨 Color Legend")
    cl1, cl2, cl3, cl4 = st.columns(4)
    with cl1:
        st.markdown(
            "<div style='background:#22543d; color:#C6F6D5; padding:12px; border-radius:8px; text-align:center; font-weight:600;'>"
            "🟢 GOOD<br><small>> 0.8N</small></div>", unsafe_allow_html=True,
        )
    with cl2:
        st.markdown(
            "<div style='background:#744210; color:#FEFCBF; padding:12px; border-radius:8px; text-align:center; font-weight:600;'>"
            "🟡 WARNING<br><small>0.35 – 0.8N</small></div>", unsafe_allow_html=True,
        )
    with cl3:
        st.markdown(
            "<div style='background:#9b2c2c; color:#FED7D7; padding:12px; border-radius:8px; text-align:center; font-weight:600;'>"
            "🔴 FAILURE<br><small>≤ 0.35N</small></div>", unsafe_allow_html=True,
        )
    with cl4:
        st.markdown(
            "<div style='background:#742a2a; color:#FED7D7; padding:12px; border-radius:8px; text-align:center; font-weight:600;'>"
            "⚫ CRITICAL<br><small>≤ 0.1N</small></div>", unsafe_allow_html=True,
        )


# ╔══════════════════════════════════════════════════════════════════╗
# ║                        PAGE: LOGS                               ║
# ╚══════════════════════════════════════════════════════════════════╝
elif page == "📜 Logs":
    st.title("📜 System Logs")
    st.caption("View pipeline execution logs")

    log_files = sorted(LOGS_DIR.glob("*.log*"), key=os.path.getctime, reverse=True)

    if not log_files:
        st.info("No logs generated yet. Process a file to begin.")
    else:
        col_log1, col_log2 = st.columns([1, 3])

        with col_log1:
            st.markdown("##### Log Files")
            selected_log = st.radio(
                "Select log file",
                options=log_files,
                format_func=lambda p: p.name,
                label_visibility="collapsed",
            )

        with col_log2:
            if selected_log:
                st.markdown(f"##### {selected_log.name}")
                try:
                    with open(selected_log, "r", encoding="utf-8") as f:
                        lines = f.readlines()

                    # Controls
                    lc1, lc2 = st.columns([1, 1])
                    with lc1:
                        num_lines = st.slider("Lines to show", 10, min(len(lines), 500), min(50, len(lines)))
                    with lc2:
                        log_filter = st.text_input("Filter logs", "", placeholder="Type to filter...")

                    display_lines = lines[-num_lines:]
                    if log_filter:
                        display_lines = [l for l in display_lines if log_filter.lower() in l.lower()]

                    log_text = "".join(display_lines)
                    st.code(log_text, language="log")
                    st.caption(f"Showing {len(display_lines)} of {len(lines)} total lines")
                except Exception as e:
                    st.error(f"Could not read log file: {e}")
